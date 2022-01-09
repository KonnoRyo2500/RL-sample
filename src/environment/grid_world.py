# 強化学習勉強用サンプルプログラム Grid World環境クラス

from copy import copy
from itertools import product

from openpyxl import load_workbook

from common.const_val import *

# Grid World環境クラス
class GridWorld:
    # コンストラクタ
    def __init__(self, config):
        self.config = self._adjust_config(config)
        self.pos = copy(self.config['initial_pos'])
        self.wall = self._read_wall()

    # 移動+報酬の取得
    def move(self, direction):
        next_pos = self._get_next_pos(direction)
        goal_pos = self.config['goal_pos']

        # 報酬の確定
        reward = None
        if not self._can_move(direction):
            # 移動できない場合は想定しない
            # エージェントは事前に移動可能な方向をget_available_directionで取得する想定
            return reward
        elif next_pos in goal_pos:
            # ゴール時
            i = goal_pos.index(next_pos)
            reward = self.config['goal_reward'][i]
        else:
            # 通常移動時
            reward = 0

        # 移動の実施
        self.pos = copy(next_pos)

        return reward

    # 盤面のリセット
    def reset(self):
        self.pos = copy(self.config['initial_pos'])

    # 移動できる方向の検索
    def get_available_direction(self):
        available_direction = []
        for name in Direction._member_names_:
            if self._can_move(Direction[name]):
                available_direction.append(Direction[name])

        if len(available_direction) == 0:
            print(self.pos)
        return available_direction

    # 現在位置の取得
    def get_pos(self):
        return self.pos

    # 設定値の取得
    def get_config(self):
        return self.config

    # 設定値を一部調整する(型の変換などを行う)
    def _adjust_config(self, config):
        config['initial_pos'] = tuple(config['initial_pos'])
        config['goal_pos'] = [tuple(p) for p in config['goal_pos']]

        return config

    # 与えられた方向から、移動先のマスを取得する
    def _get_next_pos(self, direction):
        pos_diff = {
            Direction.Up: (0, -1),
            Direction.Down: (0, 1),
            Direction.Left: (-1, 0),
            Direction.Right: (1, 0),
        }[direction]

        next_pos = tuple([i + d for i, d in zip(self.pos, pos_diff)])

        return next_pos

    # 指定した方向に移動できるかどうか判定する
    def _can_move(self, direction):
        next_pos = self._get_next_pos(direction)

        # すでにゴールにいる場合は移動しない
        if self.pos in self.config['goal_pos']:
            return False
        next_x, next_y = next_pos

        # 範囲内の判定
        x_is_in_grid = (0 <= next_x < self.config['width'])
        y_is_in_grid = (0 <= next_y < self.config['height'])
        if (not x_is_in_grid) or (not y_is_in_grid):
            return False

        # 壁の判定
        wall = self.wall[self.pos]
        hit_wall = ((direction == Direction.Up) and (wall[0])) or \
                   ((direction == Direction.Down) and (wall[1])) or \
                   ((direction == Direction.Left) and (wall[2])) or \
                   ((direction == Direction.Right) and (wall[3]))
        if hit_wall:
            return False

        return True

    # 盤面を記述したExcelシートから、壁の情報を読み込む
    def _read_wall(self):
        book = load_workbook(self.config['grid_file'])
        # シート名が変更されていても読み込めるようにしておく。
        sheet = book._sheets[0]
        wall = {}

        # 各マスの上下左右にある壁を取得する。
        # 壁は、Excelシートの罫線で記述される。
        # マスは、A1セルから記述されるものとする。
        width, height = self.config['width'], self.config['height']
        for x, y in product(range(width), range(height)):
            c, r = x + 1, y + 1
            cell = sheet.cell(column=c, row=r)
            wall_around_cell = (
                cell.border.top.style is not None,
                cell.border.bottom.style is not None,
                cell.border.left.style is not None,
                cell.border.right.style is not None,
            )
            wall[(x, y)] = wall_around_cell

        return wall
