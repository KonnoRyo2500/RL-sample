# 強化学習勉強用サンプルプログラム Grid World環境クラス

from copy import copy
from itertools import product
from enum import Enum, auto
import os.path as op

from openpyxl import load_workbook

from environment.base.env_base import EnvironmentBase
from common.dirs import ENV_CONFIG_DIR
from common.const_val import Environment


# 移動方向
class Direction(Enum):
    Up = auto()  # 上
    Down = auto()  # 下
    Left = auto()  # 左
    Right = auto()  # 右


# Grid World環境クラス
class GridWorld(EnvironmentBase):
    # コンストラクタ
    def __init__(self):
        super().__init__(Environment.GridWorld.value)
        self.config = self._adjust_config(self.config)
        self.state = copy(self.config['initial_pos'])
        self.wall = self._read_wall()
        self.step_count = 0

    # 環境の行動空間を取得
    def get_action_space(self):
        return [d for d in Direction]

    # 現在選択可能な行動を取得
    def get_available_actions(self):
        return [d for d in Direction if self._can_move(d)]

    # 指定された行動を実行し、報酬を得る
    def exec_action(self, action):
        next_pos = self._get_next_pos(action)
        goal_pos = self.config['goal_pos']

        # 報酬の確定
        if not self._can_move(action):
            # 「行動できない」=「行動しても何も起こらない」とみなす
            return 0
        elif next_pos in goal_pos:
            # ゴール時
            i = goal_pos.index(next_pos)
            reward = self.config['goal_reward'][i]
        elif (self.step_count >= self.config['step_limit'] - 1) and (next_pos not in goal_pos):
            # ステップ数上限に達してもゴールできなかったとき
            reward = self.config['fail_reward']
        else:
            # 通常移動時
            reward = 0

        # 移動の実施
        self.state = copy(next_pos)
        self.step_count += 1

        return reward

    # 環境の状態空間を取得
    # 状態が多すぎて(もしくは無限に存在して)取得できない場合はNoneを返す
    def get_state_space(self):
        width, height = self.config['width'], self.config['height']
        return [(x, y) for x, y in product(range(width), range(height))]

    # 現在の状態を取得
    def get_state(self):
        return self.state

    # 現在の状態が終端状態かどうかを返す
    def is_terminal_state(self):
        is_goaled = (self.state in self.config['goal_pos'])
        is_failed = (self.step_count >= self.config['step_limit']) and (not is_goaled)
        return is_goaled or is_failed

    # 環境をリセットする
    def reset(self):
        self.state = copy(self.config['initial_pos'])
        self.step_count = 0

    # 設定値を一部調整する(型の変換などを行う)
    def _adjust_config(self, config):
        config['initial_pos'] = tuple(config['initial_pos'])
        config['goal_pos'] = [tuple(p) for p in config['goal_pos']]

        return config

    # 与えられた方向から、移動先のマスを取得する
    def _get_next_pos(self, direction):
        pos_diff = {
            Direction.Up: (0, -1),
            Direction.Down: (0, 1),
            Direction.Left: (-1, 0),
            Direction.Right: (1, 0),
        }[direction]

        next_pos = tuple([i + d for i, d in zip(self.state, pos_diff)])

        return next_pos

    # 指定した方向に移動できるかどうか判定する
    def _can_move(self, direction):
        next_pos = self._get_next_pos(direction)

        # すでにゴールにいる場合は移動しない
        if self.state in self.config['goal_pos']:
            return False
        next_x, next_y = next_pos

        # 範囲内の判定
        x_is_in_grid = (0 <= next_x < self.config['width'])
        y_is_in_grid = (0 <= next_y < self.config['height'])
        if (not x_is_in_grid) or (not y_is_in_grid):
            return False

        # 壁の判定
        wall = self.wall[self.state]
        hit_wall = ((direction == Direction.Up) and (wall[0])) or \
                   ((direction == Direction.Down) and (wall[1])) or \
                   ((direction == Direction.Left) and (wall[2])) or \
                   ((direction == Direction.Right) and (wall[3]))
        if hit_wall:
            return False

        return True

    # 盤面を記述したExcelシートから、壁の情報を読み込む
    def _read_wall(self):
        path = op.join(ENV_CONFIG_DIR('grid_world'), self.config['grid_file'])
        book = load_workbook(path)
        # シート名が変更されていても読み込めるようにしておく。
        sheet = book._sheets[0]
        wall = {}

        # 各マスの上下左右にある壁を取得する。
        # 壁は、Excelシートの罫線で記述される。
        # マスは、A1セルから記述されるものとする。
        width, height = self.config['width'], self.config['height']
        for x, y in product(range(width), range(height)):
            c, r = x + 1, y + 1
            cell = sheet.cell(column=c, row=r)
            wall_around_cell = (
                cell.border.top.style is not None,
                cell.border.bottom.style is not None,
                cell.border.left.style is not None,
                cell.border.right.style is not None,
            )
            wall[(x, y)] = wall_around_cell

        return wall
